#Credits: Tech with Tim channel. 
from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment


data = {
    "Leonardo": {
        "math": 78,
        "english": 98,
        "history": 88,
        "science": 92
    },
    "Maria": {
        "math": 65,
        "english": 85,
        "history": 90,
        "science": 70
    },
    "João": {
        "math": 90,
        "english": 95,
        "history": 100,
        "science": 80,
    },
    "Pedro": {
        "math": 70,
        "english": 80,
        "history": 70,
        "science": 60
    },  
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Leonardo'].keys())
ws.append(headings)


for person in data: 
    grades = list(data[person].values())
    ws.append([person] + grades)

#Calculo das medias de notas
for col in range(2, len(data['Leonardo']) + 2): #Adicionamos 2 pois estamos começando a contar a partir da coluna 2. 
    char = get_column_letter(col)
    ws[char+"6"] = f"=SUM({char + '2'}:{char + '5'})/{len(data)}" #B2 até B5, etc 
    

#Formatação de células
for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold = True, color="00990D")


wb.save("example_grades.xlsx")