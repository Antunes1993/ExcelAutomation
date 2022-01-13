from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter

wb = load_workbook(filename = 'Book1.xlsx')
ws = wb.active

#Exibindo valores de células
'''
print(ws['A1'].value)
print(ws['B1'].value)
print(ws['C1'].value)
'''
#Alterando valores de células
# Obs: Precisa usar o comando de save para salvar as alterações.
# Para salvar, a planilha precisa estar fechada.
ws['A2'] = 'Leonardo'
wb.save('Book1.xlsx')   


#Criando uma nova planilha
wb.create_sheet(title = 'Nova Planilha')
wb.save('Book1.xlsx')


#Criando um novo workbook 
wb = Workbook()
ws  = wb.active
ws.title = 'Nova Planilha - Data'

#Inserindo valores na planilha manualmente
ws['A1'] = 'Data'
ws['B1'] = 'Data1'

#Inserindo valores em linhas com listas
ws.append(['Data2', 'Data3', 'Data4'])
wb.save("NovaPlanilha.xlsx")


#Loop em diferentes células na planilha
wb = load_workbook('Book1.xlsx')
ws = wb.active

for row in range(1, 8):
    for col in range(1, 6):
        #print(ws.cell(row = row, column = col).value)

        #Obtendo o caracter que representa a célula
        char = chr(col + 64) #Forma 01 para exibir "A, B, C, etc"
        char = get_column_letter(col) #Forma 02 para exibir "A, B, C, etc" - Pega um inteiro de 1 a 26 e traz o caracter correspondente
        print(ws[char + str(row)].value)

#Merge e Unmerge Cells
wb = load_workbook(filename = 'NovaPlanilha.xlsx')
ws = wb.active
ws.merge_cells('A1:C1')     #Merge de A1 até C1
ws.unmerge_cells('A1:C1')   #Para desfazer o merge

ws.merge_cells('A1:D2')     #Merge de A1 até D2
ws.unmerge_cells('A1:D2')   #Para desfazer o merge
wb.save("NovaPlanilha.xlsx")


#Inserindo e deletando linhas
ws.insert_rows(7)
ws.delete_rows(7)

#Inserindo e deletando colunas
ws.insert_cols(7)
ws.delete_cols(7)


#Mover células 
ws.move_range("A1:C1", rows = 1, cols = 1)
wb.save("NovaPlanilha.xlsx")
